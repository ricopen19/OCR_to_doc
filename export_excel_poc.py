from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from lxml import html
from PIL import Image, ImageFilter, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

from plain_text import to_plain_text


@dataclass
class TableCell:
    row: int
    col: int
    text: str = ""
    row_span: int = 1
    col_span: int = 1
    box: list[int] | None = None


SYMBOL_WHITELIST = "○〇◯◎△▲▼×✕xX□■●"
SYMBOL_NORMALIZE = {
    "✕": "×",
    "x": "×",
    "X": "×",
    "〇": "○",
    "◯": "○",
    "O": "○",
    "o": "○",
    "0": "○",
    "口": "□",
    "■": "■",
    "□": "□",
    "●": "●",
}

_FILENAME_INVALID_RE = re.compile(r'[<>:"/\\\\|?*\\x00-\\x1F]')


def _crop_inner(img: Image.Image, box: list[int], *, margin_ratio: float = 0.18) -> Image.Image:
    x0, y0, x1, y1 = (int(box[0]), int(box[1]), int(box[2]), int(box[3]))
    w = max(1, x1 - x0)
    h = max(1, y1 - y0)
    mx = max(1, int(round(w * margin_ratio)))
    my = max(1, int(round(h * margin_ratio)))
    left = x0 + mx
    top = y0 + my
    right = x1 - mx
    bottom = y1 - my
    if right <= left or bottom <= top:
        # margin が大きすぎる場合は元の box で切る
        left, top, right, bottom = x0, y0, x1, y1
    return img.crop((left, top, right, bottom))


def _black_ratio(bw: Image.Image) -> float:
    hist = bw.histogram()
    if not hist:
        return 0.0
    total = sum(hist)
    if total <= 0:
        return 0.0
    # 二値化済み (0 or 255) を前提
    black = hist[0]
    return black / total


def _binarize_for_symbol(img: Image.Image, *, threshold: int = 210) -> Image.Image:
    gray = ImageOps.autocontrast(img.convert("L"))
    return gray.point(lambda p: 0 if p < threshold else 255, mode="L")


def _looks_like_ring(bw: Image.Image) -> bool:
    """○ に近い輪郭（外周が濃く、中心が白寄り）を雑に判定する。"""

    w, h = bw.size
    if w < 8 or h < 8:
        return False

    overall = _black_ratio(bw)
    if overall < 0.01 or overall > 0.25:
        return False

    cx0 = int(w * 0.30)
    cy0 = int(h * 0.30)
    cx1 = int(w * 0.70)
    cy1 = int(h * 0.70)
    center = bw.crop((cx0, cy0, cx1, cy1))
    center_ratio = _black_ratio(center)

    border = max(1, int(min(w, h) * 0.15))
    edge_parts = [
        bw.crop((0, 0, w, border)),
        bw.crop((0, h - border, w, h)),
        bw.crop((0, 0, border, h)),
        bw.crop((w - border, 0, w, h)),
    ]
    edge_pixels = sum(sum(p.histogram()) for p in edge_parts)
    if edge_pixels <= 0:
        return False
    edge_black = sum(p.histogram()[0] for p in edge_parts)
    edge_ratio = edge_black / edge_pixels

    if edge_ratio <= center_ratio * 1.8:
        return False

    # 左右・上下のバランス（円形っぽいか）
    left = bw.crop((0, 0, w // 2, h))
    right = bw.crop((w // 2, 0, w, h))
    top = bw.crop((0, 0, w, h // 2))
    bottom = bw.crop((0, h // 2, w, h))
    lr = _black_ratio(left) / max(_black_ratio(right), 1e-6)
    tb = _black_ratio(top) / max(_black_ratio(bottom), 1e-6)
    if not (0.6 <= lr <= 1.6 and 0.6 <= tb <= 1.6):
        return False

    return True


def _detect_outline_symbol(bw: Image.Image) -> str | None:
    """輪郭だけの記号（○/□）を簡易検出する（OpenCV なし）。"""

    # 線が薄いと輪郭が途切れるので少し太らせる（黒=0 が広がる）
    bw = bw.filter(ImageFilter.MinFilter(3)).filter(ImageFilter.MinFilter(3))

    # 小さめにして計算量を抑える
    target = bw.resize((64, 64), Image.Resampling.NEAREST)
    pixels = list(target.getdata())
    grid = [[0] * 64 for _ in range(64)]
    black_points: list[tuple[int, int]] = []
    for i, v in enumerate(pixels):
        x = i % 64
        y = i // 64
        is_black = 1 if v == 0 else 0
        grid[y][x] = is_black
        if is_black:
            black_points.append((x, y))

    if not black_points:
        return None

    # 白(0)の外部領域を flood fill し、内部の白があれば「穴あり」
    from collections import deque

    visited = [[False] * 64 for _ in range(64)]
    q: deque[tuple[int, int]] = deque()
    for x in range(64):
        for y in (0, 63):
            if grid[y][x] == 0 and not visited[y][x]:
                visited[y][x] = True
                q.append((x, y))
    for y in range(64):
        for x in (0, 63):
            if grid[y][x] == 0 and not visited[y][x]:
                visited[y][x] = True
                q.append((x, y))

    while q:
        x, y = q.popleft()
        for dx, dy in ((1, 0), (-1, 0), (0, 1), (0, -1)):
            nx, ny = x + dx, y + dy
            if nx < 0 or nx >= 64 or ny < 0 or ny >= 64:
                continue
            if visited[ny][nx]:
                continue
            if grid[ny][nx] == 1:  # black
                continue
            visited[ny][nx] = True
            q.append((nx, ny))

    hole_pixels = 0
    for y in range(64):
        for x in range(64):
            if grid[y][x] == 0 and not visited[y][x]:
                hole_pixels += 1

    # 穴が小さすぎるのはノイズ扱い
    if hole_pixels < 20:
        return None

    xs = [p[0] for p in black_points]
    ys = [p[1] for p in black_points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    bw_w = max(1, max_x - min_x + 1)
    bw_h = max(1, max_y - min_y + 1)
    patch = max(2, int(round(min(bw_w, bw_h) * 0.18)))

    def corner_ratio(x0: int, y0: int) -> float:
        black = 0
        total = 0
        for yy in range(y0, min(64, y0 + patch)):
            for xx in range(x0, min(64, x0 + patch)):
                total += 1
                if grid[yy][xx] == 1:
                    black += 1
        return black / max(total, 1)

    corners = [
        corner_ratio(min_x, min_y),
        corner_ratio(max(0, max_x - patch + 1), min_y),
        corner_ratio(min_x, max(0, max_y - patch + 1)),
        corner_ratio(max(0, max_x - patch + 1), max(0, max_y - patch + 1)),
    ]
    avg_corner = sum(corners) / len(corners)

    # 角が黒いほど四角形寄り。円は角が抜けやすい。
    return "□" if avg_corner >= 0.06 else "○"


def _ocr_symbol_tesseract(img: Image.Image) -> str | None:
    try:
        import pytesseract  # type: ignore
    except Exception:
        return None

    # 記号は小さいので拡大して渡す
    scale = 4
    resized = img.resize((img.width * scale, img.height * scale), Image.Resampling.NEAREST)
    bw = _binarize_for_symbol(resized)

    # まずは OCR
    config = f"--psm 10 --oem 1 -c tessedit_char_whitelist={SYMBOL_WHITELIST}"
    text = pytesseract.image_to_string(bw, lang="jpn+eng", config=config) or ""
    text = text.strip()
    if text:
        ch = next((c for c in text if not c.isspace()), "")
        ch = SYMBOL_NORMALIZE.get(ch, ch)
        if ch in SYMBOL_WHITELIST or ch in SYMBOL_NORMALIZE.values():
            return ch

    # OCR が空でも、輪郭記号（○/□）は画像処理で補完
    detected = _detect_outline_symbol(_binarize_for_symbol(img, threshold=235))
    if detected:
        return detected
    return None


def fill_empty_cells_with_symbols(
    page_image_path: Path,
    tables: list[list[TableCell]],
    *,
    min_black_ratio: float = 0.012,
) -> int:
    """JSON tables の空セルに対して、画像から記号を補完する。"""

    if not page_image_path.exists():
        return 0

    filled = 0
    with Image.open(page_image_path) as page_img:
        page_img = ImageOps.exif_transpose(page_img)
        for cells in tables:
            for cell in cells:
                if cell.text.strip():
                    continue
                box = getattr(cell, "box", None)
                if not box:
                    continue
                crop = _crop_inner(page_img, box)
                bw = _binarize_for_symbol(crop)
                if _black_ratio(bw) < min_black_ratio:
                    continue
                symbol = _ocr_symbol_tesseract(crop)
                if symbol:
                    cell.text = symbol
                    filled += 1
    return filled


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert YomiToku outputs to Excel")
    parser.add_argument("input", type=Path, help="入力ファイル (json/csv/html)")
    parser.add_argument("output", type=Path, help="出力する xlsx パス")
    parser.add_argument(
        "--format",
        choices=["json", "csv", "html"],
        required=True,
        help="入力フォーマット",
    )
    parser.add_argument(
        "--csv-tables-only",
        action="store_true",
        help="CSV の場合、1 列だけの段落ブロックを除外し、複数列ブロックのみを表として扱う",
    )
    parser.add_argument(
        "--sheet-prefix",
        default="table",
        help="シート名のプレフィックス (default: table)",
    )
    parser.add_argument(
        "--meta",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="メタ情報シートを出力するか (default: true)",
    )
    parser.add_argument(
        "--review-columns",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="表の右隣にレビュー用列（確認メモ/ステータス）を追加する",
    )
    parser.add_argument(
        "--auto-format",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="数値/日付/百分率を自動判定して書式設定する (default: true)",
    )
    parser.add_argument(
        "--excel-mode",
        choices=["layout", "table"],
        default="layout",
        help="xlsx の出力モード。layout=レイアウト優先（結合あり）、table=結合解除してテーブル化 (default: layout)",
    )
    return parser.parse_args()


def load_tables_from_json(
    path: Path,
    *,
    page_image_path: Path | None = None,
    enable_symbol_fallback: bool = False,
) -> list[list[TableCell]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    tables = data.get("tables") or []
    result: list[list[TableCell]] = []
    for table in tables:
        cells: list[TableCell] = []
        for cell in table.get("cells", []):
            cells.append(
                TableCell(
                    row=int(cell.get("row", 1)),
                    col=int(cell.get("col", 1)),
                    text=(cell.get("contents") or "").strip(),
                    row_span=max(1, int(cell.get("row_span", 1))),
                    col_span=max(1, int(cell.get("col_span", 1))),
                    box=cell.get("box"),
                )
            )
        if cells:
            result.append(cells)
    if not result:
        raise ValueError("JSON 内に table/cell が見つかりません")

    if enable_symbol_fallback and page_image_path is not None:
        fill_empty_cells_with_symbols(page_image_path, result)
    return result


def load_tables_from_csv(path: Path, *, tables_only: bool = False) -> list[list[TableCell]]:
    tables: list[list[TableCell]] = []
    current: list[list[str]] = []
    with path.open(encoding="utf-8") as fp:
        reader = csv.reader(fp)
        for row in reader:
            if not any(cell.strip() for cell in row):
                if current:
                    tables.append(current)
                    current = []
                continue
            current.append(row)
    if current:
        tables.append(current)

    if tables_only:
        # 1 列だけのブロック（段落とみなす）を除外し、複数列を含むブロックだけ残す
        filtered = []
        for block in tables:
            max_cols = max((len(row) for row in block), default=0)
            if max_cols <= 1:
                continue
            filtered.append(block)
        tables = filtered

    result: list[list[TableCell]] = []
    for table in tables:
        cells: list[TableCell] = []
        for r_idx, row in enumerate(table, start=1):
            for c_idx, value in enumerate(row, start=1):
                cells.append(TableCell(row=r_idx, col=c_idx, text=value.strip(), box=None))
        if cells:
            result.append(cells)

    if not result:
        raise ValueError("CSV 内に表データが見つかりません")
    return result


def load_tables_from_html(path: Path) -> list[list[TableCell]]:
    content = path.read_text(encoding="utf-8")
    doc = html.fromstring(content or "<html></html>")
    table_elements = doc.findall(".//table")
    if not table_elements:
        raise ValueError("HTML 内に <table> が見つかりません")

    tables: list[list[TableCell]] = []
    for table_el in table_elements:
        cells: list[TableCell] = []
        occupied: dict[tuple[int, int], bool] = {}
        row_index = 0
        for tr in table_el.findall(".//tr"):
            row_index += 1
            col_index = 1
            # スパンで埋まっている座標をスキップ
            while (row_index, col_index) in occupied:
                col_index += 1
            for cell_el in tr.findall(".//th") + tr.findall(".//td"):
                while (row_index, col_index) in occupied:
                    col_index += 1
                text = cell_el.text_content().strip()
                row_span = int(cell_el.get("rowspan", "1") or 1)
                col_span = int(cell_el.get("colspan", "1") or 1)
                cells.append(
                    TableCell(
                        row=row_index,
                        col=col_index,
                        text=text,
                        row_span=max(1, row_span),
                        col_span=max(1, col_span),
                        box=None,
                    )
                )
                for r in range(row_index, row_index + row_span):
                    for c in range(col_index, col_index + col_span):
                        occupied[(r, c)] = True
                col_index += col_span
        if cells:
            tables.append(cells)

    return tables


def auto_adjust_columns(ws) -> None:
    """列幅を簡易調整しつつ折り返しを活かす。

    - 短文は固定幅広め（読みやすさ優先）
    - 長文は幅を抑え、wrap_text で折り返す
    """

    for column_cells in ws.columns:
        values = [len(str(cell.value)) for cell in column_cells if cell.value]
        if not values:
            continue
        letter = get_column_letter(column_cells[0].column)
        max_len = max(values)
        if max_len <= 8:
            width = 10
        elif max_len <= 20:
            width = 18
        elif max_len <= 40:
            width = 28
        else:
            width = 40  # 長文は折り返し前提
        ws.column_dimensions[letter].width = width


def split_text_to_paragraphs(text: str) -> list[str]:
    """空行区切りで段落配列にする（Markdown/プレーンテキスト想定）。"""

    paragraphs: list[str] = []
    buffer: list[str] = []
    for line in (text or "").splitlines():
        if line.strip():
            buffer.append(line.rstrip())
            continue
        if buffer:
            paragraphs.append("\n".join(buffer).strip())
            buffer = []
    if buffer:
        paragraphs.append("\n".join(buffer).strip())
    return paragraphs


def write_text_to_workbook(
    paragraphs: list[str],
    *,
    sheet_name: str = "本文",
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(["本文"])
    for para in paragraphs:
        ws.append([para])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 80
    align = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_col=1):
        row[0].alignment = align

    return wb


PERCENT_RE = re.compile(r"^\s*-?\d+[\d,\.]*\s*%\s*$")
INT_RE = re.compile(r"^\s*-?\d{1,3}(?:,\d{3})*\s*$")
FLOAT_RE = re.compile(r"^\s*-?\d*[\.,]?\d+\s*$")
DATE_RE = [
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y.%m.%d",
]


def apply_auto_format(target_cell, text: str, *, enable: bool) -> bool:
    """値を数値/日付/百分率に変換し、number_format を付与する。

    戻り値: True=値をセット済み、False=元の文字列を使うこと。
    """

    if not enable or not isinstance(text, str):
        return False

    raw = text.strip()
    if not raw:
        return False

    # percentage
    if PERCENT_RE.match(raw):
        num_str = raw.replace("%", "").replace(",", "")
        num_str = num_str.replace("％", "")
        try:
            value = float(num_str) / 100.0
        except ValueError:
            return False
        target_cell.value = value
        target_cell.number_format = "0.0%" if "." in num_str else "0%"
        return True

    # integer
    if INT_RE.match(raw):
        try:
            value = int(raw.replace(",", ""))
        except ValueError:
            return False
        target_cell.value = value
        target_cell.number_format = "#,##0"
        return True

    # float
    if FLOAT_RE.match(raw):
        try:
            value = float(raw.replace(",", ""))
        except ValueError:
            return False
        target_cell.value = value
        target_cell.number_format = "#,##0.00"
        return True

    # date
    for fmt in DATE_RE:
        try:
            dt = datetime.strptime(raw, fmt).date()
            target_cell.value = dt
            target_cell.number_format = "yyyy/mm/dd"
            return True
        except ValueError:
            continue

    return False


def _sanitize_excel_sheet_title(name: str, *, fallback: str) -> str:
    """Excel シート名として安全な文字列にする（最大 31 文字）。"""

    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in (name or "").strip())
    cleaned = cleaned.strip("'").strip()
    if not cleaned:
        cleaned = fallback
    return cleaned[:31]


def _make_unique_name(base: str, used: set[str], *, max_len: int) -> str:
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 10_000):
        suffix = f"_{i}"
        candidate = (base[: max_len - len(suffix)] + suffix) if len(base) + len(suffix) > max_len else base + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise RuntimeError("failed to allocate unique name")


def _sanitize_table_display_name(name: str, *, fallback: str) -> str:
    """Excel テーブル名（ListObject）として安全な文字列にする。"""

    cleaned = re.sub(r"[^0-9A-Za-z_]", "_", (name or "").strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned:
        cleaned = fallback
    if not re.match(r"^[A-Za-z_]", cleaned):
        cleaned = f"_{cleaned}"
    return cleaned[:255]


def _build_owner_and_value_grids(cells: list[TableCell]) -> tuple[list[list[int | None]], list[list[str]]]:
    max_row = 0
    max_col = 0
    for cell in cells:
        max_row = max(max_row, cell.row + cell.row_span - 1)
        max_col = max(max_col, cell.col + cell.col_span - 1)
    max_row = max(max_row, 1)
    max_col = max(max_col, 1)

    owner: list[list[int | None]] = [[None for _ in range(max_col)] for _ in range(max_row)]
    values: list[list[str]] = [["" for _ in range(max_col)] for _ in range(max_row)]
    for idx, cell in enumerate(cells):
        for r in range(cell.row - 1, cell.row - 1 + cell.row_span):
            if r < 0 or r >= max_row:
                continue
            for c in range(cell.col - 1, cell.col - 1 + cell.col_span):
                if c < 0 or c >= max_col:
                    continue
                # 既に埋まっている場合は先勝ち（JSON 側の重なりを吸収）
                if owner[r][c] is None:
                    owner[r][c] = idx
                if not values[r][c]:
                    values[r][c] = cell.text
                else:
                    # 値が既にある場合でも、同一セルの埋め戻しは許容
                    pass
    return owner, values


def _row_signature(owner_row: list[int | None]) -> tuple[int, ...]:
    """行のセル境界（結合含む）を signature 化する。"""

    if not owner_row:
        return tuple()

    def normalize(v: int | None, col_index: int):
        # None（JSON 上でセル定義が欠ける）を連結扱いにしないよう、列ごとに別IDにする
        return v if v is not None else ("__none__", col_index)

    sig: list[int] = []
    last = normalize(owner_row[0], 0)
    run = 1
    for col_index, x in enumerate(owner_row[1:], start=1):
        nx = normalize(x, col_index)
        if nx == last:
            run += 1
            continue
        sig.append(run)
        last = nx
        run = 1
    sig.append(run)
    return tuple(sig)


def _is_blank_row(values_row: list[str]) -> bool:
    return not any((v or "").strip() for v in values_row)


def _extract_table_segments_by_structure(
    cells: list[TableCell],
) -> list[tuple[list[int], list[int], int, int]]:
    """セル構造の変化に基づき「ヘッダー行群 + データ行群」のセグメントへ分割する。

    戻り値: [(header_rows, data_rows, max_row, max_col), ...]
    header_rows/data_rows は 1 起点の行番号配列。
    """

    owner, values = _build_owner_and_value_grids(cells)
    max_row = len(owner)
    max_col = len(owner[0]) if owner else 0

    signatures: list[tuple[int, ...] | None] = []
    for r in range(max_row):
        if _is_blank_row(values[r]):
            signatures.append(None)
        else:
            signatures.append(_row_signature(owner[r]))

    def row_has_symbol(row_values: list[str]) -> bool:
        for v in row_values:
            t = (v or "").strip()
            if not t:
                continue
            t = SYMBOL_NORMALIZE.get(t, t)
            if t in SYMBOL_WHITELIST or t in SYMBOL_NORMALIZE.values():
                return True
        return False

    # まずは「行の構造（signature）」が連続する run にまとめる（空行は区切り）。
    runs: list[list[int]] = []
    current: list[int] = []
    cur_sig: tuple[int, ...] | None = None
    for i, sig in enumerate(signatures, start=1):
        if sig is None:
            if current:
                runs.append(current)
                current = []
                cur_sig = None
            continue
        if cur_sig is None or sig != cur_sig:
            if current:
                runs.append(current)
            current = [i]
            cur_sig = sig
        else:
            current.append(i)
    if current:
        runs.append(current)

    if not runs:
        return []

    has_any_symbol = any(
        row_has_symbol(values[r - 1]) for r in range(1, max_row + 1) if signatures[r - 1] is not None
    )

    segments: list[tuple[list[int], list[int], int, int]] = []
    pending_header: list[int] = []

    if has_any_symbol:
        for rows in runs:
            first_symbol = next((r for r in rows if row_has_symbol(values[r - 1])), None)
            if first_symbol is None:
                pending_header.extend(rows)
                continue
            header_rows = pending_header + [r for r in rows if r < first_symbol]
            data_rows = [r for r in rows if r >= first_symbol]
            if data_rows:
                segments.append((header_rows, data_rows, max_row, max_col))
            pending_header = []
        return segments

    # 記号が無い表向けのフォールバック: 最長 run をデータ扱いにして、間の run をヘッダー扱いにする。
    max_len = max(len(r) for r in runs)
    for rows in runs:
        if len(rows) == max_len:
            segments.append((pending_header, rows, max_row, max_col))
            pending_header = []
        else:
            pending_header.extend(rows)
    return segments


def _merge_header_rows(
    values: list[list[str]],
    header_rows: list[int],
    *,
    max_col: int,
) -> list[str]:
    header: list[str] = []
    for c in range(1, max_col + 1):
        parts: list[str] = []
        for r in header_rows:
            text = (values[r - 1][c - 1] or "").strip()
            if not text:
                continue
            if not parts or parts[-1] != text:
                parts.append(text)
        merged = " / ".join(parts).strip()
        header.append(merged)

    # 空ヘッダー補完 + 重複回避
    used: set[str] = set()
    out: list[str] = []
    for idx, raw in enumerate(header, start=1):
        base = raw or f"col_{idx}"
        base = base.replace("\n", " ").strip() or f"col_{idx}"
        unique = _make_unique_name(base, used, max_len=255)
        out.append(unique)
    return out


def _detect_table_name(values: list[list[str]], data_rows: list[int]) -> str:
    """表名（シート名候補）を推定する。基本はデータ行の 1 列目。"""

    for r in data_rows:
        v = (values[r - 1][0] or "").strip()
        if v:
            return v
    return ""


def _trim_max_col(values: list[list[str]], rows: list[int], *, max_col: int) -> int:
    last = 0
    for r in rows:
        row = values[r - 1]
        for c in range(max_col, 0, -1):
            if (row[c - 1] or "").strip():
                last = max(last, c)
                break
    return max(1, last) if last else max_col


def write_tables_to_workbook_table_mode(
    tables: list[list[TableCell]],
    *,
    sheet_prefix: str,
    review_columns: bool,
    auto_format: bool,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    align = Alignment(vertical="top", wrap_text=True)
    align_center = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    used_sheet_names: set[str] = set()
    used_table_names: set[str] = set()

    segment_index = 0
    for table_cells in tables:
        if not table_cells:
            continue
        _, values = _build_owner_and_value_grids(table_cells)
        max_col = len(values[0]) if values else 0
        segments = _extract_table_segments_by_structure(table_cells)
        if not segments:
            continue

        for header_rows, data_rows, _, _ in segments:
            if not data_rows:
                continue
            segment_index += 1
            header_rows = header_rows or []
            block_rows = header_rows + data_rows
            block_max_col = _trim_max_col(values, block_rows, max_col=max_col)

            table_name_raw = _detect_table_name(values, data_rows) or f"{sheet_prefix}_{segment_index}"
            sheet_base = _sanitize_excel_sheet_title(table_name_raw, fallback=f"{sheet_prefix}_{segment_index}")
            sheet_name = _make_unique_name(sheet_base, used_sheet_names, max_len=31)

            ws = wb.create_sheet(title=sheet_name)

            header = _merge_header_rows(values, header_rows, max_col=block_max_col) if header_rows else [
                f"col_{i}" for i in range(1, block_max_col + 1)
            ]

            # ヘッダー
            for c, text in enumerate(header, start=1):
                cell = ws.cell(row=1, column=c)
                cell.value = to_plain_text(text)
                cell.alignment = align_center

            # データ
            out_row = 2
            for r in data_rows:
                for c in range(1, block_max_col + 1):
                    raw = to_plain_text(values[r - 1][c - 1])
                    target = ws.cell(row=out_row, column=c)
                    if not apply_auto_format(target, raw, enable=auto_format):
                        target.value = raw
                    target.alignment = align
                out_row += 1

            max_row_out = out_row - 1
            table_max_col = block_max_col

            # レビュー列（表から2列あけてステータスのみ、○/× ドロップダウン）
            status_col_opt: int | None = None
            if review_columns and max_row_out >= 1:
                status_col = table_max_col + 3
                ws.cell(row=1, column=status_col).value = "ステータス"
                dv = DataValidation(type="list", formula1='"○,×"', allow_blank=True)
                ws.add_data_validation(dv)
                status_range = f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{max_row_out}"
                dv.add(status_range)
                ws.cell(row=1, column=status_col).alignment = align_center
                for rr in range(2, max_row_out + 1):
                    ws.cell(row=rr, column=status_col).alignment = align_center
                status_col_opt = status_col

            # 罫線: 表本体とステータス列だけ
            border_cols = list(range(1, table_max_col + 1))
            if status_col_opt:
                border_cols.append(status_col_opt)
            for rr in range(1, max_row_out + 1):
                for cc in border_cols:
                    ws.cell(row=rr, column=cc).border = thin

            ws.freeze_panes = "A2"

            # Excel テーブル化（必ず作る）
            if max_row_out >= 1 and table_max_col >= 1:
                ref = f"A1:{get_column_letter(table_max_col)}{max_row_out}"
                display_name_base = _sanitize_table_display_name(
                    f"{sheet_prefix}_{sheet_name}_tbl",
                    fallback=f"{sheet_prefix}_{segment_index}_tbl",
                )
                display_name = _make_unique_name(display_name_base, used_table_names, max_len=255)
                table = Table(
                    displayName=display_name,
                    ref=ref,
                    tableStyleInfo=TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    ),
                )
                ws.add_table(table)

            auto_adjust_columns(ws)

    return wb


def _sanitize_filename(name: str, *, fallback: str) -> str:
    cleaned = _FILENAME_INVALID_RE.sub("_", (name or "").strip())
    cleaned = cleaned.strip().strip(".")
    if not cleaned:
        cleaned = fallback
    # Windows reserved names
    reserved = {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        *(f"COM{i}" for i in range(1, 10)),
        *(f"LPT{i}" for i in range(1, 10)),
    }
    if cleaned.upper() in reserved:
        cleaned = f"_{cleaned}"
    return cleaned[:120]


def write_tables_to_csv_files(
    tables: list[list[TableCell]],
    *,
    output_dir: Path,
    base_name: str,
    excel_mode: str = "table",
) -> list[Path]:
    """JSON tables から CSV を出力する。

    - table: 結合解除＋構造変化で分割（実務向け）
    - layout: JSON の表グリッドをそのまま CSV 化（結合は表現できないため値は埋める）
    """

    output_dir.mkdir(parents=True, exist_ok=True)
    used: set[str] = set()
    outputs: list[Path] = []
    segment_index = 0

    for table_cells in tables:
        if not table_cells:
            continue

        _, values = _build_owner_and_value_grids(table_cells)
        max_row = len(values)
        max_col = len(values[0]) if values else 0

        if excel_mode == "layout":
            segment_index += 1
            # 末尾の空列/空行を落とす（見た目のノイズ削減）
            rows = list(range(1, max_row + 1))
            block_max_col = _trim_max_col(values, rows, max_col=max_col)
            last_row = 0
            for r in range(max_row, 0, -1):
                if any((v or "").strip() for v in values[r - 1][:block_max_col]):
                    last_row = r
                    break
            last_row = max(1, last_row) if max_row else 1

            base = f"{base_name}__table_{segment_index:02d}"
            unique = _make_unique_name(base, used, max_len=200)
            out_path = output_dir / f"{unique}.csv"
            with out_path.open("w", encoding="utf-8", newline="") as fp:
                writer = csv.writer(fp)
                for r in range(1, last_row + 1):
                    row = [
                        to_plain_text(values[r - 1][c - 1])
                        for c in range(1, block_max_col + 1)
                    ]
                    writer.writerow(row)
            outputs.append(out_path)
            continue

        segments = _extract_table_segments_by_structure(table_cells)
        if not segments:
            continue

        for header_rows, data_rows, _, _ in segments:
            if not data_rows:
                continue
            segment_index += 1
            block_rows = (header_rows or []) + data_rows
            block_max_col = _trim_max_col(values, block_rows, max_col=max_col)

            header = (
                _merge_header_rows(values, header_rows, max_col=block_max_col)
                if header_rows
                else [f"col_{i}" for i in range(1, block_max_col + 1)]
            )

            table_name_raw = _detect_table_name(values, data_rows) or f"{base_name}_{segment_index}"
            file_label = _sanitize_filename(table_name_raw, fallback=f"{base_name}_{segment_index}")
            base = f"{base_name}__{file_label}"
            unique = _make_unique_name(base, used, max_len=200)
            out_path = output_dir / f"{unique}.csv"

            with out_path.open("w", encoding="utf-8", newline="") as fp:
                writer = csv.writer(fp)
                writer.writerow([to_plain_text(h) for h in header])
                for r in data_rows:
                    row = [
                        to_plain_text(values[r - 1][c - 1])
                        for c in range(1, block_max_col + 1)
                    ]
                    writer.writerow(row)

            outputs.append(out_path)

    return outputs


def write_tables_to_workbook(
    tables: list[list[TableCell]],
    *,
    sheet_prefix: str,
    review_columns: bool,
    auto_format: bool,
    excel_mode: str = "layout",
) -> Workbook:
    if excel_mode == "table":
        return write_tables_to_workbook_table_mode(
            tables,
            sheet_prefix=sheet_prefix,
            review_columns=review_columns,
            auto_format=auto_format,
        )

    wb = Workbook()
    wb.remove(wb.active)
    align = Alignment(vertical="top", wrap_text=True)
    align_center = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for idx, cells in enumerate(tables, start=1):
        ws = wb.create_sheet(title=f"{sheet_prefix}_{idx}")
        max_row = 0
        max_col = 0
        for cell in cells:
            target = ws.cell(row=cell.row, column=cell.col)
            raw = to_plain_text(cell.text)
            if not apply_auto_format(target, raw, enable=auto_format):
                target.value = raw
            target.alignment = align
            if cell.row_span > 1 or cell.col_span > 1:
                ws.merge_cells(
                    start_row=cell.row,
                    start_column=cell.col,
                    end_row=cell.row + cell.row_span - 1,
                    end_column=cell.col + cell.col_span - 1,
                )
            max_row = max(max_row, cell.row + cell.row_span - 1)
            max_col = max(max_col, cell.col + cell.col_span - 1)

        table_max_col = max_col  # Table 範囲は元の表のみ

        # レビュー列（表から2列あけてステータスのみ、○/× ドロップダウン）
        if review_columns and max_row > 0:
            status_col = table_max_col + 3  # 2 列スペースを空ける
            ws.cell(row=1, column=status_col).value = "ステータス"
            dv = DataValidation(type="list", formula1='"○,×"', allow_blank=True)
            ws.add_data_validation(dv)
            status_range = f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{max_row}"
            dv.add(status_range)
            ws.cell(row=1, column=status_col).alignment = align_center
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=status_col).alignment = align_center
            max_col = status_col
        status_col_opt = status_col if review_columns and max_row > 0 else None


        # 罫線: 表本体 (1..table_max_col) とステータス列だけ。間の空列は塗らない。
        border_cols = list(range(1, table_max_col + 1))
        if status_col_opt:
            border_cols.append(status_col_opt)

        for r in range(1, max_row + 1):
            for c in border_cols:
                ws.cell(row=r, column=c).border = thin

        # Excel テーブル化（先頭行をヘッダーと見なす）
        # Excel の Table は結合セルを含められないため、結合セルが存在する場合はスキップする。
        has_merges = bool(ws.merged_cells.ranges)
        if not has_merges and max_row >= 1 and table_max_col >= 1:
            ref = f"A1:{get_column_letter(table_max_col)}{max_row}"
            table = Table(
                displayName=f"{sheet_prefix}_{idx}_tbl",
                ref=ref,
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                ),
            )
            ws.add_table(table)

        auto_adjust_columns(ws)
    return wb


def add_meta_sheet(wb: Workbook, *, args: argparse.Namespace, sheet_names: list[str]) -> None:
    ws = wb.create_sheet(title="meta", index=0)
    ws.append(["項目", "値"])
    ws.append(["入力ファイル", str(args.input)])
    ws.append(["入力フォーマット", args.format])
    ws.append(["出力ファイル", str(args.output)])
    ws.append(["生成日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["シート数", len(sheet_names)])

    ws.append([])
    ws.append(["シートリンク", ""])
    for name in sheet_names:
        row = ws.max_row + 1
        cell = ws.cell(row=row, column=2)
        cell.value = name
        cell.hyperlink = f"#{name}!A1"
        cell.style = "Hyperlink"

    auto_adjust_columns(ws)


def main() -> None:
    args = parse_args()
    if args.format == "json":
        tables = load_tables_from_json(args.input)
    elif args.format == "csv":
        tables = load_tables_from_csv(args.input, tables_only=args.csv_tables_only)
    else:
        tables = load_tables_from_html(args.input)

    workbook = write_tables_to_workbook(
        tables,
        sheet_prefix=args.sheet_prefix,
        review_columns=args.review_columns,
        auto_format=args.auto_format,
        excel_mode=args.excel_mode,
    )
    sheet_names = workbook.sheetnames.copy()
    if args.meta:
        add_meta_sheet(workbook, args=args, sheet_names=sheet_names)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Saved {args.output}")


if __name__ == "__main__":
    main()
